import os

import pandas as pd
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists="replace")

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


if __name__ == '__main__':
    output_file = 'gathered_sheet.xlsx'
    files = [
        ('amulet', 'export/amulet_balances_long.csv'),
        ('pauldron', 'export/pauldron_balances_long.csv'),
        ('gun', 'export/gun_balances_long.csv'),
        ('melee', 'export/melee_balances_long.csv'),
        ('ring', 'export/ring_balances_long.csv'),
        ('shield', 'export/shield_balances_long.csv'),
        ('spell', 'export/spell_balances_long.csv'),
    ]

    for row in files:
        print("generating {} page".format(row[0]))
        read_file = pd.read_csv(row[1])
        append_df_to_excel(
            output_file,
            read_file,
            sheet_name=row[0],
            startrow=0,
            truncate_sheet=True,
            index=None,
            header=True,
        )
